import pygame
import sys
import random
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

lebar_display, tinggi_display = 640, 480
satu_kotak = 40
total_lebar_kotak = lebar_display // satu_kotak
total_tinggi_kotak = tinggi_display // satu_kotak
file_excel = "snake_leaderboard.xlsx"

pygame.init()
pygame.mixer.init()

screen = pygame.display.set_mode((lebar_display, tinggi_display))
pygame.display.set_caption("SNAKE MUKBANG - Alexa, Aulia, Leon")
frame_rate = pygame.time.Clock()
font_dasar = pygame.font.SysFont("Minecraft", 24)

awal = os.path.dirname(os.path.abspath(__file__))
assets_tujuan = os.path.join(awal, "my_assets")

def load_img(name, scale=None):
    path = os.path.join(assets_tujuan, name)
    img = pygame.image.load(path).convert_alpha()
    if scale:
        img = pygame.transform.scale(img, scale)
    return img

background_img = load_img("background.png", (lebar_display, tinggi_display))

#main_menu_img sudah dipakai di HOME sebagai gambar besar (200x200)
#kita buat versi kecil buat tombol "home" di gameplay
main_menu_img = load_img("main_menu.png", (200, 200))
home_btn_img = load_img("main_menu.png", (60, 60))  # tombol di samping pause

start_img = load_img("start.png", (100, 100))
quit_img = load_img("quit.png", (100, 100))
pause_img = load_img("pause.png", (60, 60))
restart_img = load_img("restart.png", (100, 100))
continue_img = load_img("continue.png", (100, 100))
food_img = load_img("food.png", (50, 50))
gameover_img = load_img("gameover.png", (200, 200))

head_imgs = {d: load_img(f"head_{d}.png", (satu_kotak, satu_kotak)) for d in ("up", "down", "left", "right")}
tail_imgs = {d: load_img(f"tail_{d}.png", (satu_kotak, satu_kotak)) for d in ("up", "down", "left", "right")}

body_straight = load_img("body_straight.png", (satu_kotak, satu_kotak))
body_turn = load_img("body_turn.png", (satu_kotak, satu_kotak))

turn_imgs = {
    f"{a}_{b}": load_img(f"{a}_{b}.png", (satu_kotak, satu_kotak))
    for a, b in [("up", "right"), ("up", "left"), ("down", "left"), ("down", "right")]
}

def init_excel():
    if not os.path.exists(file_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "Leaderboard"
        ws.append(["Nama", "Score", "Timestamp", "Mode"])
        wb.save(file_excel)
        return

    wb = load_workbook(file_excel)
    ws = wb.active
    header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if header != ["Nama", "Score", "Timestamp", "Mode"]:
        data = list(ws.iter_rows(values_only=True))
        rows = data[1:] if len(data) > 1 else []
        ws.delete_rows(1, ws.max_row)
        ws.append(["Nama", "Score", "Timestamp", "Mode"])
        for r in rows:
            nama = r[0] if len(r) > 0 else ""
            score = int(r[1]) if len(r) > 1 and r[1] is not None else 0
            ts = r[2] if len(r) > 2 and r[2] is not None else ""
            mode = r[3] if len(r) > 3 and r[3] is not None else "Unknown"
            ws.append([nama, score, ts, mode])
        wb.save(file_excel)

def clear_leaderboard():
    init_excel()
    wb = load_workbook(file_excel)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(file_excel)

def save_score(name, score, mode):
    wb = load_workbook(file_excel)
    ws = wb.active

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([name, int(score), ts, mode])

    data = list(ws.iter_rows(values_only=True))
    header, rows = data[0], data[1:]

    def parse_ts(row):
        try:
            return datetime.strptime(row[2], "%Y-%m-%d %H:%M:%S")
        except:
            return datetime.min

    rows.sort(key=lambda r: (int(r[1]) if r[1] is not None else 0, parse_ts(r)), reverse=True)

    ws.delete_rows(1, ws.max_row)
    ws.append(list(header))
    for r in rows:
        ws.append(list(r))

    wb.save(file_excel)

def get_all_scores():
    if not os.path.exists(file_excel):
        return []
    wb = load_workbook(file_excel)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cleaned = []
    for r in rows:
        if not r or r[0] is None:
            continue
        nama = str(r[0])
        score = int(r[1]) if len(r) > 1 and r[1] is not None else 0
        ts = r[2] if len(r) > 2 and r[2] is not None else ""
        mode = r[3] if len(r) > 3 and r[3] is not None else ""
        cleaned.append((nama, score, ts, mode))
    cleaned.sort(key=lambda x: x[1], reverse=True)
    return cleaned

init_excel()

pygame.mixer.music.load(os.path.join(assets_tujuan, "background_music.wav"))
pygame.mixer.music.play(-1)

def draw_button(rect, text, font, bg_color, border_color, text_color):
    pygame.draw.rect(screen, bg_color, rect, border_radius=10)
    pygame.draw.rect(screen, border_color, rect, 2, border_radius=10)
    surf = font.render(text, True, text_color)
    screen.blit(surf, surf.get_rect(center=rect.center))

class SnakeGame:
    def __init__(self):
        self.eat_sound = pygame.mixer.Sound(os.path.join(assets_tujuan, "pou_eating.wav"))
        self.gameover_sound = pygame.mixer.Sound(os.path.join(assets_tujuan, "fairytale_gameover.wav"))

        self.mode = "Normal"
        self.speed_map = {"Normal": 5, "Hard": 10}

        self.ask_name = True
        self.player_name = ""

        self.state = "HOME"  # HOME / LEADERBOARD / PLAYING

        self.saved_this_round = False

        # leaderboard page
        self.lb_page = 0
        self.lb_per_page = 8

        # confirm clear
        self.confirm_clear = False

        # rects gameplay buttons
        self.pause_rect = pygame.Rect(570, 10, 60, 60)

        # HOME button (sebelah pause)
        # taruh di kiri pause, jadi (570-70, 10) = (500, 10)
        self.home_btn_rect = pygame.Rect(500, 10, 60, 60)

        self.continue_rect = pygame.Rect(0, 0, 0, 0)
        self.restart_rect = pygame.Rect(0, 0, 0, 0)
        self.quit_rect = pygame.Rect(0, 0, 0, 0)

        # home menu buttons
        self.home_start_rect = pygame.Rect(lebar_display // 2 - 120, tinggi_display // 2 + 80, 100, 100)
        self.home_quit_rect = pygame.Rect(lebar_display // 2 + 20, tinggi_display // 2 + 80, 100, 100)
        self.home_leader_rect = pygame.Rect(lebar_display // 2 - 100, tinggi_display // 2 + 200, 200, 45)

        # leaderboard page buttons
        self.lb_back_rect = pygame.Rect(20, 20, 110, 40)
        self.lb_clear_rect = pygame.Rect(lebar_display - 150, 20, 130, 40)
        self.lb_prev_rect = pygame.Rect(20, tinggi_display - 60, 110, 40)
        self.lb_next_rect = pygame.Rect(lebar_display - 130, tinggi_display - 60, 110, 40)

        # confirm dialog buttons (center)
        self.confirm_yes_rect = pygame.Rect(lebar_display // 2 - 130, tinggi_display // 2 + 50, 110, 45)
        self.confirm_no_rect = pygame.Rect(lebar_display // 2 + 20, tinggi_display // 2 + 50, 110, 45)

        self.reset()

    def go_home(self):
        # balik ke menu utama (HOME)
        self.state = "HOME"
        self.paused = False
        self.game_over = False
        # tidak reset nama & mode, biar tetap
        # gameplay akan di-reset lagi pas start

    def reset(self):
        mid_x = total_lebar_kotak // 2
        mid_y = total_tinggi_kotak // 2
        self.snake = [(mid_x, mid_y), (mid_x, mid_y + 1), (mid_x, mid_y + 2)]
        self.direction = (0, -1)
        self.spawn_food()
        self.score = 0
        self.game_over = False
        self.paused = False
        self.saved_this_round = False

    def spawn_food(self):
        while True:
            pos = (random.randint(0, total_lebar_kotak - 1), random.randint(0, total_tinggi_kotak - 1))
            if pos not in self.snake:
                self.food = pos
                return

    def get_direction(self, p1, p2):
        x1, y1 = p1
        x2, y2 = p2
        if x1 == x2:
            return "down" if y2 > y1 else "up"
        return "right" if x2 > x1 else "left"

    def move(self):
        if self.state != "PLAYING":
            return
        if self.paused or self.game_over:
            return

        head_x, head_y = self.snake[0]
        dir_x, dir_y = self.direction
        new_head = (head_x + dir_x, head_y + dir_y)

        if not (0 <= new_head[0] < total_lebar_kotak and 0 <= new_head[1] < total_tinggi_kotak) or new_head in self.snake:
            self.game_over = True
            self.gameover_sound.play()
            if not self.saved_this_round:
                save_score(self.player_name or "Player", self.score, self.mode)
                self.saved_this_round = True
            return

        self.snake.insert(0, new_head)

        if new_head == self.food:
            self.score += 1
            self.eat_sound.play()
            self.spawn_food()
        else:
            self.snake.pop()

    def draw_input_name(self):
        screen.blit(background_img, (0, 0))
        title = font_dasar.render("Masukkan Nama Anda:", True, pygame.Color("#B30000"))
        screen.blit(title, (lebar_display // 2 - 130, tinggi_display // 2 - 80))

        pygame.draw.rect(screen, pygame.Color("#B30000"),
                         (lebar_display // 2 - 150, tinggi_display // 2 - 30, 300, 40), 2)

        name_surf = font_dasar.render(self.player_name, True, pygame.Color("#B30000"))
        screen.blit(name_surf, (lebar_display // 2 - 140, tinggi_display // 2 - 20))

        hint = font_dasar.render("Tekan ENTER untuk lanjut!", True, pygame.Color("#B30000"))
        screen.blit(hint, (lebar_display // 2 - 165, tinggi_display // 2 + 40))

    def draw_home(self):
        screen.blit(background_img, (0, 0))
        overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
        overlay.fill((250, 0, 0, 100))
        screen.blit(overlay, (0, 0))

        title_font = pygame.font.SysFont("Minecraft", 48)
        title_surf = title_font.render("SNAKE MUKBANG", True, pygame.Color("#FFDEDE"))
        screen.blit(title_surf, title_surf.get_rect(center=(lebar_display // 2, tinggi_display // 2 - 150)))

        sub_font = pygame.font.SysFont("Minecraft", 20)
        sub_surf = sub_font.render("ALEXA, AULIA, LEON", True, pygame.Color("#FFDEDE"))
        screen.blit(sub_surf, sub_surf.get_rect(center=(lebar_display // 2, tinggi_display // 2 - 96)))

        screen.blit(main_menu_img, main_menu_img.get_rect(center=(lebar_display // 2, tinggi_display // 2)))
        screen.blit(start_img, (self.home_start_rect.x, self.home_start_rect.y))
        screen.blit(quit_img, (self.home_quit_rect.x, self.home_quit_rect.y))

        draw_button(
            self.home_leader_rect,
            "LEADERBOARD",
            pygame.font.SysFont("Minecraft", 20),
            pygame.Color("#FFDEDE"),
            pygame.Color("#B30000"),
            pygame.Color("#B30000"),
        )

        hint_surf = pygame.font.SysFont("Minecraft", 18).render(
            f"MODE: {self.mode} (Tekan 1=Normal, 2=Hard)", True, pygame.Color("#FFDEDE")
        )
        screen.blit(hint_surf, (20, tinggi_display - 70))

    def draw_confirm_dialog(self):
        overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
        overlay.fill((0, 0, 0, 160))
        screen.blit(overlay, (0, 0))

        box = pygame.Rect(lebar_display // 2 - 220, tinggi_display // 2 - 110, 440, 220)
        pygame.draw.rect(screen, pygame.Color("#FFDEDE"), box, border_radius=12)
        pygame.draw.rect(screen, pygame.Color("#B30000"), box, 3, border_radius=12)

        title = pygame.font.SysFont("Minecraft", 26).render("CLEAR LEADERBOARD?", True, pygame.Color("#B30000"))
        screen.blit(title, title.get_rect(center=(lebar_display // 2, tinggi_display // 2 - 40)))

        msg = pygame.font.SysFont("Minecraft", 18).render("Semua data akan dihapus. Yakin?", True, pygame.Color("#B30000"))
        screen.blit(msg, msg.get_rect(center=(lebar_display // 2, tinggi_display // 2 - 5)))

        draw_button(self.confirm_yes_rect, "YES", pygame.font.SysFont("Minecraft", 20),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))
        draw_button(self.confirm_no_rect, "NO", pygame.font.SysFont("Minecraft", 20),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))

    def draw_leaderboard_page(self):
        screen.blit(background_img, (0, 0))
        overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
        overlay.fill((250, 0, 0, 120))
        screen.blit(overlay, (0, 0))

        header_font = pygame.font.SysFont("Minecraft", 40)
        header = header_font.render("LEADERBOARD", True, pygame.Color("#FFDEDE"))
        screen.blit(header, header.get_rect(center=(lebar_display // 2, 60)))

        draw_button(self.lb_back_rect, "BACK", pygame.font.SysFont("Minecraft", 20),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))
        draw_button(self.lb_clear_rect, "CLEAR", pygame.font.SysFont("Minecraft", 20),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))

        table_font = pygame.font.SysFont("Minecraft", 18)
        col = pygame.Color("#FFDEDE")
        screen.blit(table_font.render("Rank", True, col), (60, 110))
        screen.blit(table_font.render("Name", True, col), (130, 110))
        screen.blit(table_font.render("Score", True, col), (330, 110))
        screen.blit(table_font.render("Mode", True, col), (420, 110))
        screen.blit(table_font.render("Time", True, col), (500, 110))
        pygame.draw.line(screen, pygame.Color("#FFDEDE"), (50, 135), (590, 135), 2)

        all_scores = get_all_scores()
        total = len(all_scores)
        per_page = self.lb_per_page
        max_page = max(0, (total - 1) // per_page)
        self.lb_page = max(0, min(self.lb_page, max_page))

        start = self.lb_page * per_page
        end = min(start + per_page, total)
        page_rows = all_scores[start:end]

        y = 150
        row_font = pygame.font.SysFont("Minecraft", 18)
        for i, (nama, score, ts, mode) in enumerate(page_rows, start=start + 1):
            nama_show = (nama[:12] + "...") if len(nama) > 15 else nama
            ts_show = ts[5:] if isinstance(ts, str) and len(ts) >= 16 else str(ts)

            screen.blit(row_font.render(str(i), True, col), (60, y))
            screen.blit(row_font.render(nama_show, True, col), (130, y))
            screen.blit(row_font.render(str(score), True, col), (340, y))
            screen.blit(row_font.render(str(mode), True, col), (420, y))
            screen.blit(row_font.render(ts_show, True, col), (500, y))
            y += 30

        info_font = pygame.font.SysFont("Minecraft", 18)
        info = info_font.render(f"Page {self.lb_page + 1}/{max_page + 1} | Total: {total}", True, pygame.Color("#FFDEDE"))
        screen.blit(info, info.get_rect(center=(lebar_display // 2, tinggi_display - 40)))

        draw_button(self.lb_prev_rect, "PREV", pygame.font.SysFont("Minecraft", 18),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))
        draw_button(self.lb_next_rect, "NEXT", pygame.font.SysFont("Minecraft", 18),
                    pygame.Color("#FFDEDE"), pygame.Color("#B30000"), pygame.Color("#B30000"))

        hint = pygame.font.SysFont("Minecraft", 16).render("Klik PREV/NEXT | ESC=Back | C=Clear", True, pygame.Color("#FFDEDE"))
        screen.blit(hint, (lebar_display - 450, tinggi_display - 25))

        if self.confirm_clear:
            self.draw_confirm_dialog()

    def draw_playing(self):
        screen.blit(background_img, (0, 0))

        screen.blit(food_img, (self.food[0] * satu_kotak, self.food[1] * satu_kotak))

        # ====== TOMBOL HOME (sebelah pause) + TOMBOL PAUSE ======
        screen.blit(home_btn_img, self.home_btn_rect.topleft)
        screen.blit(pause_img, self.pause_rect.topleft)

        snake = self.snake

        if len(snake) > 1:
            head_dir = self.get_direction(snake[0], snake[1])
        else:
            x, y = self.direction
            head_dir = "up" if y < 0 else "down" if y > 0 else "right" if x > 0 else "left"
        screen.blit(head_imgs[head_dir], (snake[0][0] * satu_kotak, snake[0][1] * satu_kotak))

        for i in range(1, len(snake) - 1):
            prev, cur, nex = snake[i - 1], snake[i], snake[i + 1]
            dir_prev = self.get_direction(cur, prev)
            dir_next = self.get_direction(cur, nex)

            if dir_prev in ("up", "down") and dir_next in ("up", "down"):
                screen.blit(body_straight, (cur[0] * satu_kotak, cur[1] * satu_kotak))
            elif dir_prev in ("left", "right") and dir_next in ("left", "right"):
                screen.blit(pygame.transform.rotate(body_straight, 90), (cur[0] * satu_kotak, cur[1] * satu_kotak))
            else:
                key1 = f"{dir_prev}_{dir_next}"
                key2 = f"{dir_next}_{dir_prev}"
                screen.blit(turn_imgs.get(key1, turn_imgs.get(key2, body_turn)),
                            (cur[0] * satu_kotak, cur[1] * satu_kotak))

        if len(snake) > 1:
            tail_dir = self.get_direction(snake[-2], snake[-1])
        else:
            x, y = self.direction
            tail_dir = "up" if y < 0 else "down" if y > 0 else "right" if x > 0 else "left"
        screen.blit(tail_imgs[tail_dir], (snake[-1][0] * satu_kotak, snake[-1][1] * satu_kotak))

        screen.blit(font_dasar.render(f"Score: {self.score}", True, pygame.Color("#B30000")), (10, 10))
        screen.blit(font_dasar.render(f"Mode: {self.mode}", True, pygame.Color("#B30000")), (10, 40))

        if self.paused:
            overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
            overlay.fill((250, 0, 67, 120))
            screen.blit(overlay, (0, 0))

            paused_surf = pygame.font.SysFont("Minecraft", 48).render("PAUSED", True, pygame.Color("#FFDEDE"))
            screen.blit(paused_surf, paused_surf.get_rect(center=(lebar_display // 2, tinggi_display // 2 - 80)))

            cx, cy = lebar_display // 2, tinggi_display // 2
            buttons = [
                (restart_img, cx - 170, cy, "restart"),
                (continue_img, cx - 50, cy, "continue"),
                (quit_img, cx + 70, cy, "quit"),
            ]
            for img, x, y, name in buttons:
                screen.blit(img, (x, y))
                setattr(self, f"{name}_rect", pygame.Rect(x, y, 100, 100))

            # hint tambahan: tekan HOME icon untuk balik menu
            hint = pygame.font.SysFont("Minecraft", 18).render("Klik icon HOME untuk balik ke menu", True, pygame.Color("#FFDEDE"))
            screen.blit(hint, hint.get_rect(center=(lebar_display // 2, tinggi_display // 2 + 140)))

        if self.game_over:
            screen.blit(gameover_img, (lebar_display // 2 - 103, tinggi_display // 2 - 150))
            over_txt = pygame.font.SysFont("Minecraft", 20).render(
                "AUTO-SAVED! SPACE=Restart | L=Leaderboard | Q=Quit",
                True, pygame.Color("#B30000")
            )
            screen.blit(over_txt, over_txt.get_rect(center=(lebar_display // 2, tinggi_display // 2 + 200)))

    def draw(self):
        if self.ask_name:
            self.draw_input_name()
            return

        if self.state == "HOME":
            self.draw_home()
        elif self.state == "LEADERBOARD":
            self.draw_leaderboard_page()
        elif self.state == "PLAYING":
            self.draw_playing()

game = SnakeGame()
running = True

while running:
    fps = game.speed_map.get(game.mode, 5) if game.state == "PLAYING" else 60
    frame_rate.tick(fps)

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
            continue

        #Input nama
        if game.ask_name and event.type == pygame.KEYDOWN:
            if event.key == pygame.K_RETURN and game.player_name:
                game.ask_name = False
                game.state = "HOME"
            elif event.key == pygame.K_BACKSPACE:
                game.player_name = game.player_name[:-1]
            elif event.unicode.isprintable() and len(game.player_name) < 12:
                game.player_name += event.unicode
            continue

        #Mouse click
        if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            pos = event.pos

            # HOME
            if game.state == "HOME":
                if game.home_start_rect.collidepoint(pos):
                    game.state = "PLAYING"
                    game.reset()
                elif game.home_quit_rect.collidepoint(pos):
                    running = False
                elif game.home_leader_rect.collidepoint(pos):
                    game.state = "LEADERBOARD"
                    game.lb_page = 0
                    game.confirm_clear = False
                continue

            # LEADERBOARD
            if game.state == "LEADERBOARD":
                if game.confirm_clear:
                    if game.confirm_yes_rect.collidepoint(pos):
                        clear_leaderboard()
                        game.lb_page = 0
                        game.confirm_clear = False
                    elif game.confirm_no_rect.collidepoint(pos):
                        game.confirm_clear = False
                    continue

                if game.lb_back_rect.collidepoint(pos):
                    game.state = "HOME"
                elif game.lb_clear_rect.collidepoint(pos):
                    game.confirm_clear = True
                elif game.lb_prev_rect.collidepoint(pos):
                    game.lb_page -= 1
                elif game.lb_next_rect.collidepoint(pos):
                    game.lb_page += 1
                continue

            # PLAYING
            if game.state == "PLAYING":
                # HOME icon button (selalu bisa, bahkan saat game over / pause)
                if game.home_btn_rect.collidepoint(pos):
                    game.go_home()
                    continue

                # Pause button bisa dipencet meski gameover
                if game.pause_rect.collidepoint(pos):
                    game.paused = not game.paused

                # tombol pause menu (tetap bisa dipakai, meski game over)
                if game.paused:
                    if game.restart_rect.collidepoint(pos):
                        game.reset()
                    elif game.quit_rect.collidepoint(pos):
                        running = False
                    elif game.continue_rect.collidepoint(pos):
                        if not game.game_over:
                            game.paused = False
                continue

        #Keyboard 
        if event.type == pygame.KEYDOWN:
            # HOME keys
            if game.state == "HOME":
                if event.key == pygame.K_1:
                    game.mode = "Normal"
                elif event.key == pygame.K_2:
                    game.mode = "Hard"
                elif event.key in (pygame.K_SPACE, pygame.K_RETURN):
                    game.state = "PLAYING"
                    game.reset()
                elif event.key == pygame.K_l:
                    game.state = "LEADERBOARD"
                    game.lb_page = 0
                    game.confirm_clear = False
                elif event.key == pygame.K_q:
                    running = False
                continue

            # LEADERBOARD keys
            if game.state == "LEADERBOARD":
                if game.confirm_clear:
                    if event.key == pygame.K_y:
                        clear_leaderboard()
                        game.lb_page = 0
                        game.confirm_clear = False
                    elif event.key in (pygame.K_n, pygame.K_ESCAPE):
                        game.confirm_clear = False
                    continue

                if event.key == pygame.K_ESCAPE:
                    game.state = "HOME"
                elif event.key == pygame.K_LEFT:
                    game.lb_page -= 1
                elif event.key == pygame.K_RIGHT:
                    game.lb_page += 1
                elif event.key == pygame.K_c:
                    game.confirm_clear = True
                continue

            # PLAYING keys
            if game.state == "PLAYING":
                # tombol keyboard untuk balik menu juga (opsional)
                if event.key == pygame.K_h:
                    game.go_home()
                    continue

                if game.game_over:
                    if event.key == pygame.K_SPACE:
                        game.reset()
                    elif event.key == pygame.K_l:
                        game.state = "LEADERBOARD"
                        game.lb_page = 0
                        game.confirm_clear = False
                    elif event.key == pygame.K_q:
                        running = False
                    continue

                if event.key == pygame.K_p:
                    game.paused = not game.paused

                if game.paused:
                    if event.key == pygame.K_r:
                        game.reset()
                    continue

                if event.key == pygame.K_UP and game.direction != (0, 1):
                    game.direction = (0, -1)
                elif event.key == pygame.K_DOWN and game.direction != (0, -1):
                    game.direction = (0, 1)
                elif event.key == pygame.K_LEFT and game.direction != (1, 0):
                    game.direction = (-1, 0)
                elif event.key == pygame.K_RIGHT and game.direction != (-1, 0):
                    game.direction = (1, 0)

    game.move()
    game.draw()
    pygame.display.flip()

pygame.quit()
sys.exit()