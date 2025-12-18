import pygame
import sys
import random
from openpyxl import Workbook, load_workbook
import os

#assets
awal = os.path.dirname(os.path.abspath(__file__)) 
assets_tujuan = os.path.join(awal, "my_assets")

# ini konfigurassinyaa
lebar_display = 640
tinggi_display = 480
satu_kotak = 40

total_lebar_kotak = lebar_display // satu_kotak
total_tinggi_kotak = tinggi_display // satu_kotak

file_excel = "snake_leaderboard.xlsx"

warna_ular_dasar = pygame.Color("#FA0043")

#dasar
pygame.init() #wajib ada jika pakai library pygame
screen = pygame.display.set_mode((lebar_display, tinggi_display)) #layar display yang selalu ada
pygame.display.set_caption("PySnake - Alexa, Aulia, Leon") #judul pojok kiri atas

frame_rate = pygame.time.Clock() #agar fps bisa diatur
font_dasar = pygame.font.SysFont("Minecraft", 24) #font dasar download Minecraft.ttf

# loading assets (pastikan file ada di folder my_assets)
background_img = pygame.image.load(os.path.join(assets_tujuan,"background.png"))


main_menu_img = pygame.image.load(os.path.join(assets_tujuan, "main_menu.png"))
start_img = pygame.image.load(os.path.join(assets_tujuan, "start.png")) 
quit_img = pygame.image.load(os.path.join(assets_tujuan, "quit.png"))

food_img = pygame.image.load(os.path.join(assets_tujuan, "food.png"))
gameover_img = pygame.image.load(os.path.join(assets_tujuan,"gameover.png"))

pause_img = pygame.image.load(os.path.join(assets_tujuan,"pause.png"))
restart_img = pygame.image.load(os.path.join(assets_tujuan, "restart.png"))
continue_img = pygame.image.load(os.path.join(assets_tujuan, "continue.png"))

#load image ular & scale sekalian
# Head 
head_imgs = {
    "up": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "head_up.png")), (satu_kotak, satu_kotak)),
    "down": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "head_down.png")), (satu_kotak, satu_kotak)),
    "left": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "head_left.png")), (satu_kotak, satu_kotak)),
    "right": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "head_right.png")), (satu_kotak, satu_kotak)),
}

# Body 
body_straight = pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "body_straight.png")), (satu_kotak, satu_kotak))
body_turn = pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "body_turn.png")), (satu_kotak, satu_kotak))

# Turn 
turn_imgs = {
    "up_right": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "up_right.png")), (satu_kotak, satu_kotak)),
    "up_left": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "up_left.png")), (satu_kotak, satu_kotak)),
    "down_left": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "down_left.png")), (satu_kotak, satu_kotak)),
    "down_right": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "down_right.png")), (satu_kotak, satu_kotak)),
}

# Tail 
tail_imgs = {
    "up": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "tail_up.png")), (satu_kotak, satu_kotak)),
    "down": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "tail_down.png")), (satu_kotak, satu_kotak)),
    "left": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "tail_left.png")), (satu_kotak, satu_kotak)),
    "right": pygame.transform.scale(pygame.image.load(os.path.join(assets_tujuan, "tail_right.png")), (satu_kotak, satu_kotak)),
}

# penempatannn / scaling
main_menu_img = pygame.transform.scale(main_menu_img, (200,200))
start_img = pygame.transform.scale(start_img, (100,100))
quit_img = pygame.transform.scale(quit_img, (100,100))

food_img = pygame.transform.scale(food_img, (50, 50))
gameover_img = pygame.transform.scale(gameover_img, (200,200))

pause_img = pygame.transform.scale(pause_img, (60,60))
restart_img = pygame.transform.scale(restart_img, (100,100))
continue_img = pygame.transform.scale(continue_img, (100,100))
#scaling background
background_img = pygame.transform.scale(background_img, (lebar_display, tinggi_display))

#membuat file excel
def init_excel():
    if not os.path.exists(file_excel):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Score"])
        wb.save(file_excel)

def save_score(name, score):
    wb = load_workbook(file_excel)
    ws = wb.active
    ws.append([name, score])
    data = list(ws.iter_rows(values_only=True))
    header, rows = data[0], data[1:]
    rows.sort(key=lambda r: r[1], reverse=True)

    ws.delete_rows(1, ws.max_row)
    ws.append(header)
    for r in rows:
        ws.append(r)

    wb.save(file_excel)

init_excel()

pygame.mixer.music.load(os.path.join('my_assets/background_music.wav'))
pygame.mixer.music.play(-1)

#kelas untuk atribut dan method
class SnakeGame:
    def __init__(self): #membuat atribut yang akan dipakai di method nanti
        #backsound makan dan gameover
        self.eat_sound = pygame.mixer.Sound('my_assets/pou_eating.wav') 
        self.gameover_sound = pygame.mixer.Sound('my_assets/fairytale_gameover.wav')

        # ini input nama (baru)
        self.ask_name = True #biar awal muncul display langsung fungsi "tanya nama" nyala 
        self.player_name = "" #variabel kosong akan terisi setlah player input nama
        
        # sekarang ada state 'started' untuk home screen
        self.started = False #biar gak langsung jalan game nya
        self.reset()
        
        # tambahan untuk tombol pause clickable
        self.pause_rect = pygame.Rect(570, 10, 50, 50) #koordinat x,y, ukuran lebar x tinggi
        self.clicked = False #biar tombol gak langsung kepencet
        
        # resume/restart/quit rect akan diupdate di draw saat pause, tapi inisialisasi agar ada atribut
        self.continue_rect = pygame.Rect(0,0,0,0) 
        self.restart_rect = pygame.Rect(0,0,0,0)
        self.back_rect = pygame.Rect(0,0,0,0)
        self.quit_rect = pygame.Rect(0,0,0,0)
        # rect untuk tombol home (start + quit)
        self.home_start_rect = pygame.Rect(lebar_display//2 - 120, tinggi_display//2 + 80, 100, 100)
        self.home_quit_rect = pygame.Rect(lebar_display//2 + 20, tinggi_display//2 + 80, 100, 100)

    def get_direction(self, p1, p2):
        """Return direction string from p1 -> p2 (p1 and p2 are (x,y) tuples)."""
        x1, y1 = p1
        x2, y2 = p2
        if x1 == x2:
            return "down" if y2 > y1 else "up"
        return "right" if x2 > x1 else "left"

    def reset(self):
        # reset hanya isi permainan; started tidak diubah di sini supaya home tetap dikontrol terpisah
        # start with length 3 so corners/tail/head appear nicely (optional)
        mid_x = total_lebar_kotak // 2
        mid_y = total_tinggi_kotak // 2
        self.snake = [(mid_x, mid_y), (mid_x, mid_y+1), (mid_x, mid_y+2)]
        self.direction = (0, -1)
        self.spawn_food()
        self.score = 0
        self.game_over = False
        self.paused = False

    def spawn_food(self):
        # ensure food doesn't spawn on snake
        while True:
            posisi_food = (random.randint(0, total_lebar_kotak - 2), random.randint(0, total_tinggi_kotak - 2))
            if posisi_food not in self.snake:
                self.food = posisi_food
                return

    def move(self):
        if self.paused or self.game_over or not self.started:
            return

        head_x, head_y = self.snake[0]
        dir_x, dir_y = self.direction
        new_head = (head_x + dir_x, head_y + dir_y)

        # Nabrak tembok
        if not (0 <= new_head[0] < total_lebar_kotak and 0 <= new_head[1] < total_tinggi_kotak):
            self.game_over = True
            self.play_gameover_sound()
            return

        # Nabrak badan
        if new_head in self.snake:
            self.game_over = True
            self.play_gameover_sound()
            return

        self.snake.insert(0, new_head)

        # Makan
        if new_head == self.food:
            self.score += 1
            self.play_eat_sound()
            self.spawn_food()
        else:
            self.snake.pop()

    # draw image
    def draw(self, screen):
        # layar input nama (paling awal) (baru)
        if self.ask_name:
            screen.blit(background_img, (0, 0))

            title = font_dasar.render("Masukkan Nama Anda:", True, pygame.Color("#B30000"))
            screen.blit(title, (lebar_display//2 - 130, tinggi_display//2 - 80))

            pygame.draw.rect(screen, pygame.Color("#B30000"),
                             (lebar_display//2 - 150, tinggi_display//2 - 30, 300, 40), 2) #rectangle untuk input nama

            name_surf = font_dasar.render(self.player_name, True, pygame.Color("#B30000")) #text input an namanya
            screen.blit(name_surf, (lebar_display//2 - 140, tinggi_display//2 - 20))

            hint = font_dasar.render("Tekan ENTER untuk lanjut!", True, pygame.Color("#B30000"))
            screen.blit(hint, (lebar_display//2 - 165, tinggi_display//2 + 40))

            return   # hentikan draw di layar input nama

        # Jika belum started -> tampilkan HOME menu
        if not self.started:
            # darkened background dengan main_menu image
            screen.blit(background_img, (0, 0))
            # overlay semi transparan untuk fokus menu
            overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
            overlay.fill((250,0,0,100))
            screen.blit(overlay, (0,0))

            # Title
            title_font = pygame.font.SysFont("Minecraft", 48)
            title_surf = title_font.render("GAME PYSNAKE AAL", True, pygame.Color("#FFDEDE"))
            title_rect = title_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 150))
            screen.blit(title_surf, title_rect)

            # Subtitle or credits (tetap tampil)
            sub_font = pygame.font.SysFont("Minecraft", 20)
            sub_surf = sub_font.render("ALEXA, AULIA, LEON", True, pygame.Color("#FFDEDE"))
            sub_rect = sub_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 96))
            screen.blit(sub_surf, sub_rect)

            # main_menu_img (opsional)
            menu_img_rect = main_menu_img.get_rect(center=(lebar_display//2, tinggi_display//2))
            screen.blit(main_menu_img, menu_img_rect)

            # Tombol Start dan Quit di bawah
            screen.blit(start_img, (self.home_start_rect.x, self.home_start_rect.y))
            screen.blit(quit_img, (self.home_quit_rect.x, self.home_quit_rect.y))

            # teks petunjuk
            hint_surf = font_dasar.render("Selamat datang di Game sederhana kami!", True, pygame.Color("#FFDEDE"))
            hint_rect = hint_surf.get_rect(center=(lebar_display//2, tinggi_display//2 + 180))
            screen.blit(hint_surf, hint_rect)

            return  # jangan gambar bagian game sampai started True

        # Jika sudah started, gambar gameplay seperti biasa
        screen.blit(background_img, (0, 0))
        # update pause_rect posisi (kalau mau dinamis nanti)
        self.pause_rect.topleft = (570, 5)
        screen.blit(pause_img, self.pause_rect.topleft) #kiri,atas

        # Food pakai gambar PNG
        fx, fy = self.food
        screen.blit(food_img, (fx * satu_kotak, fy * satu_kotak))

        # INI BAGIAN ULARNYAAA
        snake = self.snake

        # ---- HEAD ----
        if len(snake) > 1:
            head = snake[0]
            neck = snake[1]
            head_dir = self.get_direction(head, neck)
        else:
            head = snake[0]
            # fallback ke arah current movement jika cuma 1 segmen
            if self.direction == (0, -1):
                head_dir = "up"
            elif self.direction == (0, 1):
                head_dir = "down"
            elif self.direction == (1, 0):
                head_dir = "right"
            else:
                head_dir = "left"

        screen.blit(head_imgs[head_dir], (head[0] * satu_kotak, head[1] * satu_kotak))

        # ---- BODY ----
        for i in range(1, len(snake)-1):
            prev = snake[i-1]
            cur = snake[i]
            nex = snake[i+1]

            dir_prev = self.get_direction(cur, prev)
            dir_next = self.get_direction(cur, nex)

            # Lurus vertikal
            if dir_prev in ("up","down") and dir_next in ("up","down"):
                screen.blit(body_straight, (cur[0]*satu_kotak, cur[1]*satu_kotak))

            # Lurus horizontal
            elif dir_prev in ("left","right") and dir_next in ("left","right"):
                straight_rot = pygame.transform.rotate(body_straight, 90)
                screen.blit(straight_rot, (cur[0]*satu_kotak, cur[1]*satu_kotak))

            # Belok
            else:
                key = f"{dir_prev}_{dir_next}"
                if key not in turn_imgs:
                    key = f"{dir_next}_{dir_prev}"
                
                img = turn_imgs.get(key, body_turn)
                screen.blit(img, (cur[0]*satu_kotak, cur[1]*satu_kotak))

        # ---- TAIL ----
        if len(snake) > 1:
            tail = snake[-1]
            before_tail = snake[-2]
            tail_dir = self.get_direction(before_tail, tail)
        else:
            tail = snake[0]
            # fallback based on movement
            if self.direction == (0, -1):
                tail_dir = "up"
            elif self.direction == (0, 1):
                tail_dir = "down"
            elif self.direction == (1, 0):
                tail_dir = "right"
            else:
                tail_dir = "left"

        screen.blit(tail_imgs[tail_dir], (tail[0]*satu_kotak, tail[1]*satu_kotak))

        # Score text (pakai font biasa)
        score_surf = font_dasar.render(f"Score: {self.score}", True, pygame.Color("#B30000"))
        screen.blit(score_surf, (10, 10))

        # Pause overlay
        if self.paused:
            overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
            overlay.fill((250, 0, 67, 100))  # RGB, transparansi
            screen.blit(overlay, (0,0))

            paused_font = pygame.font.SysFont("Minecraft", 48)
            paused_surf = paused_font.render("PAUSED", True, pygame.Color("#FFDEDE"))
            paused_rect = paused_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 80))
            screen.blit(paused_surf, paused_rect)

            # Tombol resume (start_img), restart, quit di tengah layar
            cx = lebar_display // 2
            cy = tinggi_display // 2

            restart_pos = (cx - 170, cy)
            continue_pos = (cx - 50, cy)
            quit_pos = (cx + 70, cy)

            screen.blit(restart_img, restart_pos)
            screen.blit(continue_img, continue_pos)
            screen.blit(quit_img, quit_pos)

            # update rects supaya bisa di-click pada main loop
            self.restart_rect = pygame.Rect(restart_pos[0], restart_pos[1], 100, 100)
            self.continue_rect = pygame.Rect(continue_pos[0], continue_pos[1], 100, 100)
            self.quit_rect = pygame.Rect(quit_pos[0], quit_pos[1], 100, 100)

        # Game over overlay + trophy
        if self.game_over:
            screen.blit(gameover_img, (lebar_display//2 - 103, tinggi_display//2 - 150))
            
    #fungsi untuk backsound makan
    def play_eat_sound(self): #################
        self.eat_sound.play()
        
    #fungsi untuk backsound gameover
    def play_gameover_sound(self):
        self.gameover_sound.play() ###################

    #fungsi untuk backgroundsound
    def play_background_sound(self):
        self.background_sound.play()

# ===============================
# MAIN
# ===============================

game = SnakeGame()

running = True
while running:
    frame_rate.tick(5)

    for event in pygame.event.get():
        # always handle quit first
        if event.type == pygame.QUIT:
            running = False
            continue

        # Jika masih di layar input nama: tangani keyboard untuk input nama saja
        if game.ask_name: #baru
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    if len(game.player_name) > 0:
                        game.ask_name = False
                elif event.key == pygame.K_BACKSPACE:
                    game.player_name = game.player_name[:-1]
                else:
                    if len(game.player_name) < 12 and event.unicode.isprintable():
                        game.player_name += event.unicode
            # jangan lanjutkan event lain bila masih input nama 
            continue

        # Handle mouse click untuk tombol pause & overlay & home
        if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            pos = event.pos

            # Jika belum started -> cek tombol home (start/quit)
            if not game.started:
                if game.home_start_rect.collidepoint(pos):
                    game.started = True
                    game.reset()
                elif game.home_quit_rect.collidepoint(pos):
                    running = False
                # skip sisa pengecekan karena belum started
                continue

            # jika sudah started:
            # klik tombol pause kecil di pojok
            if game.pause_rect.collidepoint(pos):
                game.paused = not game.paused

            # jika sedang pause, cek tombol overlay (resume / restart / quit)
            if game.paused:
                if game.continue_rect.width == 0:
                    cx = lebar_display // 2
                    cy = tinggi_display // 2
                    resume_rect = pygame.Rect(cx - 50, cy, 100, 100)
                    restart_rect = pygame.Rect(cx - 170, cy, 100, 100)
                    quit_rect = pygame.Rect(cx + 70, cy, 100, 100)
                else:
                    resume_rect = game.continue_rect
                    restart_rect = game.restart_rect
                    quit_rect = game.quit_rect

                if resume_rect.collidepoint(pos):
                    game.paused = False
                elif restart_rect.collidepoint(pos):
                    game.reset()
                elif quit_rect.collidepoint(pos):
                    running = False

        # keyboard events untuk gameplay (arah, pause, restart, dsb)
        if event.type == pygame.KEYDOWN:
            # Jika belum started: SPACE atau Enter untuk start, Q untuk quit
            if not game.started:
                if event.key == pygame.K_SPACE or event.key == pygame.K_RETURN:
                    game.started = True
                    game.reset()
                elif event.key == pygame.K_q:
                    running = False
                continue  # lewati pengaturan arah sampai game dimulai

            # Kontrol arah
            if event.key == pygame.K_UP and game.direction != (0, 1):
                game.direction = (0, -1)
            elif event.key == pygame.K_DOWN and game.direction != (0, -1):
                game.direction = (0, 1)
            elif event.key == pygame.K_LEFT and game.direction != (1, 0):
                game.direction = (-1, 0)
            elif event.key == pygame.K_RIGHT and game.direction != (-1, 0):
                game.direction = (1, 0)

            # Pause
            elif event.key == pygame.K_p:
                game.paused = not game.paused

            # Restart
            elif event.key == pygame.K_r:
                game.reset()

            # Save score ketika game over (pakai nama yang diinput), pencet huruf s
            if event.key == pygame.K_s and game.game_over:
                save_score(game.player_name or "Player", game.score)

            # Start kembali ketika game over
            if event.key == pygame.K_SPACE and game.game_over:
                game.reset()

    game.move()
    game.draw(screen)
    pygame.display.flip()

pygame.quit()
sys.exit()