import pygame
import sys
import random
import os
from openpyxl import Workbook, load_workbook

# ===============================
# PATH & KONFIGURASI DASAR
# ===============================
awal_dirname = os.path.dirname(os.path.abspath(__file__))
tujuan_dirname = os.path.join(awal_dirname, "my_assets")

lebar_display = 640
tinggi_display = 480
satu_kotak = 50
total_lebar_kotak = lebar_display // satu_kotak
total_tinggi_kotak = tinggi_display // satu_kotak

file_excel = "snake_leaderboard.xlsx"
warna_ular = pygame.Color("#FA0043")

# ===============================
# INIT PYGAME
# ===============================
pygame.init()
screen = pygame.display.set_mode((lebar_display, tinggi_display))
pygame.display.set_caption(" - Alexa, Aulia, Leon")
frame_rate = pygame.time.Clock()
font = pygame.font.SysFont("Minecraft", 24)

# ===============================
# LOADING & SCALING ASSETS
# ===============================
def load_and_scale(filename, size):
    path = os.path.join(tujuan_dirname, filename)
    return pygame.transform.scale(pygame.image.load(path), size)

background_img = load_and_scale("background.png", (lebar_display, tinggi_display))
food_img = load_and_scale("food.png", (satu_kotak, satu_kotak))
gameover_img = load_and_scale("gameover.png", (200, 200))
pause_img = load_and_scale("pause.png", (60, 60))
quit_img = load_and_scale("quit.png", (100, 100)) 
restart_img = load_and_scale("restart.png", (100, 100))
start_img = load_and_scale("start.png", (100, 100))
trophy_img = load_and_scale("trophy.png", (200, 200))
main_menu_img = load_and_scale("main_menu.png", (200, 200))
continue_img = load_and_scale("continue.png", (100,100))
back_img = load_and_scale("back.png", (100,100))
head_up_img = load_and_scale("head_up.png", (satu_kotak,satu_kotak))
head_right_img = load_and_scale("head_right.png", (satu_kotak,satu_kotak))
head_left_img = load_and_scale("head_left.png", (satu_kotak,satu_kotak))
head_down_img = load_and_scale("head_down.png", (satu_kotak,satu_kotak))
tail_up_img = load_and_scale("tail_up.png", (satu_kotak,satu_kotak))
tail_right_img = load_and_scale("tail_right.png", (satu_kotak,satu_kotak))
tail_left_img = load_and_scale("tail_left.png", (satu_kotak,satu_kotak))
tail_down_img = load_and_scale("tail_down.png", (satu_kotak,satu_kotak))
up_left_img = load_and_scale("up_left.png", (satu_kotak, satu_kotak))
up_right_img = load_and_scale("up_right.png", (satu_kotak, satu_kotak))
down_img_left_img = load_and_scale("down_left.png", (satu_kotak, satu_kotak))
down_right_img = load_and_scale("down_right.png", (satu_kotak, satu_kotak))

# ===============================
# LEADERBOARD FUNCTIONS
# ===============================
def init_excel():
    if not os.path.exists(file_excel):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Score"])
        wb.save(file_excel)

def save_score(name, score):
    wb = load_workbook(file_excel)
    ws = wb.active
    ws.append([name, score])
    wb.save(file_excel)

init_excel()

# ===============================
# GAME LOGIC CLASS
# ===============================
class SnakeGame:
    def __init__(self):
        # STATE HOME / START
        self.started = False
        self.reset()

        # PAUSE BUTTONS
        # RECT HOME MENU
        self.home_start_rect = pygame.Rect(lebar_display//2 - 120, tinggi_display//2 + 80, 100, 100)
        self.home_quit_rect = pygame.Rect(lebar_display//2 + 20, tinggi_display//2 + 80, 100, 100)
        self.pause_rect = pygame.Rect(570, 10, 50, 50)
        self.clicked = False

        # RECT UNTUK PAUSE MENU
        self.resume_rect = pygame.Rect(0, 0, 0, 0)
        self.restart_rect = pygame.Rect(0, 0, 0, 0)
        self.quit_rect = pygame.Rect(0, 0, 0, 0)
        self.back_rect = pygame.Rect(0, 0, 0, 0)


    def reset(self):
        self.snake = [(total_lebar_kotak // 2, total_tinggi_kotak // 2)]
        self.direction = (0, -1)
        self.spawn_food()
        self.score = 0
        self.game_over = False
        self.paused = False

    def spawn_food(self):
        self.food = (random.randint(0, total_lebar_kotak - 1), random.randint(0, total_tinggi_kotak - 1))

    def move(self):
        if self.paused or self.game_over or not self.started:
            return

        head_x, head_y = self.snake[0]
        dir_x, dir_y = self.direction
        new_head = (head_x + dir_x, head_y + dir_y)

        # Tabrakan dengan dinding
        if not (0 <= new_head[0] < total_lebar_kotak and 0 <= new_head[1] < total_tinggi_kotak):
            self.game_over = True
            return

        # Tabrakan dengan badan
        if new_head in self.snake:
            self.game_over = True
            return

        self.snake.insert(0, new_head)

        # Makan makanan
        if new_head == self.food:
            self.score += 1
            self.spawn_food()
        else:
            self.snake.pop()

    # ===============================
    # DRAW GAME
    # ===============================
    def draw(self, screen):
        # HOME MENU
        if not self.started:
            screen.blit(background_img, (0, 0))
            overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
            overlay.fill((250, 0, 0, 100))
            screen.blit(overlay, (0, 0))

            # Title & Subtitle
            title_font = pygame.font.SysFont("Minecraft", 48)
            title_surf = title_font.render("Ular Nokia", False, pygame.Color("#FAC0C0"))
            screen.blit(title_surf, title_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 150)))

            sub_font = pygame.font.SysFont("Minecraft", 20)
            sub_surf = sub_font.render("Alexa, Aulia, Leon", True, pygame.Color("#FAC0C0"))
            screen.blit(sub_surf, sub_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 96)))

            # Main menu & tombol Start/Quit
            screen.blit(main_menu_img, main_menu_img.get_rect(center=(lebar_display//2, tinggi_display//2)))
            screen.blit(start_img, (self.home_start_rect.x, self.home_start_rect.y))
            screen.blit(quit_img, (self.home_quit_rect.x, self.home_quit_rect.y))

            # Petunjuk
            hint_surf = font.render("Selamat datang di Game sederhana kami!", True, pygame.Color("#FAC0C0"))
            screen.blit(hint_surf, hint_surf.get_rect(center=(lebar_display//2, tinggi_display//2 + 180)))
            return

        # GAMEPLAY
        screen.blit(background_img, (0, 0))
        screen.blit(pause_img, self.pause_rect.topleft)

        # Food
        fx, fy = self.food
        screen.blit(food_img, (fx * satu_kotak, fy * satu_kotak))

        # Snake
        for x, y in self.snake:
            pygame.draw.rect(screen, warna_ular, (x * satu_kotak, y * satu_kotak, satu_kotak, satu_kotak))

        # Score
        score_surf = font.render(f"Score: {self.score}", True, pygame.Color("#FA0043"))
        screen.blit(score_surf, (10, 10))

        # PAUSE MENU
        if self.paused:
            overlay = pygame.Surface((lebar_display, tinggi_display), pygame.SRCALPHA)
            overlay.fill((250, 0, 67, 100))
            screen.blit(overlay, (0, 0))

            paused_font = pygame.font.SysFont("Minecraft", 48)
            paused_surf = paused_font.render("PAUSED", True, pygame.Color("#FAC0C0"))
            screen.blit(paused_surf, paused_surf.get_rect(center=(lebar_display//2, tinggi_display//2 - 80)))

            # Tombol Resume, Restart, Quit
            cx, cy = lebar_display // 2, tinggi_display // 2
            positions = {
                "restart": (cx - 170, cy + 20),   # + karena dari tinggi jendela dibagi 2 terus digeser ke bawah
                "resume": (cx -50, cy + 20),           # - karena dari lebar jendela dibagi 2 lalu digeser ke kiri
                "quit": (cx + 70, cy + 20)            # tanda minus x itu geser ke kiri, plus x ke kanan
                                                      # tanda minus y itu geser ke atas, plus y geser ke bawah
            }

            screen.blit(restart_img, positions["restart"])
            screen.blit(continue_img, positions["resume"])
            screen.blit(quit_img, positions["quit"])

            # Update rects
            self.restart_rect = pygame.Rect(*positions["restart"], 100, 100)
            self.resume_rect = pygame.Rect(*positions["resume"], 100, 100)
            self.quit_rect = pygame.Rect(*positions["quit"], 100, 100)

        # GAME OVER
        if self.game_over:
            screen.blit(gameover_img, (lebar_display//2 - 102, tinggi_display//2 - 150))   # x - itu ke kiri, x + ke kanan, y - ke atas, y + ke bawah
            #screen.blit(trophy_img, (WINDOW_WIDTH//2 - 100, WINDOW_HEIGHT//2 - 130))

# MAIN LOOP
game = SnakeGame()
running = True

#frame rate
while running:
    frame_rate.tick(3) #GANTI FRAME RATE

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

        # ===============================
        #mouse click handler
        if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            pos = event.pos

            # Home menu
            if not game.started:
                if game.home_start_rect.collidepoint(pos):
                    game.started = True
                    game.reset()
                elif game.home_quit_rect.collidepoint(pos):
                    running = False
                continue

            # Pause button
            if game.pause_rect.collidepoint(pos):
                game.paused = not game.paused

            # Pause overlay buttons
            if game.paused:
                resume_rect = game.resume_rect if game.resume_rect.width else pygame.Rect(lebar_display//2 - 50, tinggi_display//2, 100, 100)
                restart_rect = game.restart_rect if game.restart_rect.width else pygame.Rect(lebar_display//2 - 170, tinggi_display//2, 100, 100)
                quit_rect = game.quit_rect if game.quit_rect.width else pygame.Rect(lebar_display//2 + 70, tinggi_display//2, 100, 100)

                if resume_rect.collidepoint(pos):
                    game.paused = False
                elif restart_rect.collidepoint(pos):
                    game.reset()
                elif quit_rect.collidepoint(pos):
                    running = False

        # ===============================
        # KEYBOARD HANDLER
        # ===============================
        if event.type == pygame.KEYDOWN:
            # Home menu keys
            if not game.started:
                if event.key in (pygame.K_SPACE, pygame.K_RETURN):
                    game.started = True
                    game.reset()
                elif event.key == pygame.K_q:
                    running = False
                continue

            # Movement
            if event.key == pygame.K_UP and game.direction != (0, 1):
                game.direction = (0, -1)
            elif event.key == pygame.K_DOWN and game.direction != (0, -1):
                game.direction = (0, 1)
            elif event.key == pygame.K_LEFT and game.direction != (1, 0):
                game.direction = (-1, 0)
            elif event.key == pygame.K_RIGHT and game.direction != (-1, 0):
                game.direction = (1, 0)

            # Pause
            elif event.key == pygame.K_p:
                game.paused = not game.paused

            # Restart
            elif event.key == pygame.K_r:
                game.reset()

            # Save score
            if event.key == pygame.K_s and game.game_over:
                save_score("", game.score)

            # Start after game over
            if event.key == pygame.K_SPACE and game.game_over:
                game.reset()

    game.move()
    game.draw(screen)
    pygame.display.flip()

pygame.quit()
sys.exit()